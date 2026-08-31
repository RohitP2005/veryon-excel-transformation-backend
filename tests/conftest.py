import io

import pytest
from fastapi.testclient import TestClient
from openpyxl import Workbook

from app.main import app
from app.models.formula_store import formula_store


@pytest.fixture()
def client() -> TestClient:
    return TestClient(app)


@pytest.fixture(autouse=True)
def _reset_formula_store():
    """The formula store is a real file on disk (module-level singleton) - keep tests isolated."""
    formula_store._write([])
    yield
    formula_store._write([])


def _build_workbook(headers: list[str], rows: list[list]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.append(headers)
    for row in rows:
        ws.append(row)
    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


@pytest.fixture()
def customer_workbook_bytes() -> bytes:
    return _build_workbook(
        headers=["ID", "Name", "Mail", "Phone Number", "Addr1", "Addr2", "Nation"],
        rows=[
            [1, "Alice", "alice@example.com", "555-1000", "1 Main St", "Suite 5", "USA"],
            [2, "Bob", "bob@example.com", "555-2000", "2 Oak Ave", "", "Canada"],
        ],
    )


@pytest.fixture()
def pricing_workbook_bytes() -> bytes:
    return _build_workbook(headers=["Price", "Quantity"], rows=[[10.0, 3], [5.5, 2]])


@pytest.fixture()
def header_offset_workbook_bytes() -> bytes:
    """Simulates a customer file with a title row before the real header row."""
    wb = Workbook()
    ws = wb.active
    ws.append(["Fleet Report - Confidential"])
    ws.append(["ID", "Name"])
    ws.append([1, "Alice"])
    ws.append([2, "Bob"])
    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


@pytest.fixture()
def multi_level_header_workbook_bytes() -> bytes:
    """Simulates a customer file with a 2-row header: a higher-order group header row (e.g.
    "Engine 1"/"Engine 2", typed once and left blank in the cells it spans) above the real
    field-name row."""
    wb = Workbook()
    ws = wb.active
    ws.append(["Engine 1", None, "Engine 2", None])
    ws.append(["TSN", "TSO", "TSN", "TSO"])
    ws.append([12530, 100, 8000, 50])
    ws.append([13000, 120, 8200, 60])
    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


@pytest.fixture()
def three_level_header_workbook_bytes() -> bytes:
    """Simulates a 3-row header to verify the higher-order chaining works recursively for any
    number of levels, not just two."""
    wb = Workbook()
    ws = wb.active
    ws.append(["Aircraft", None, None, None])
    ws.append(["Engine 1", None, "Engine 2", None])
    ws.append(["TSN", "TSO", "TSN", "TSO"])
    ws.append([12530, 100, 8000, 50])
    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


@pytest.fixture()
def merged_header_workbook_bytes() -> bytes:
    """Simulates a customer file where a header cell (e.g. "TSA") is merged across two columns
    for visual spacing, even though only one of the two columns actually holds data beneath it -
    unmerging must not surface this as two separate "TSA" columns."""
    wb = Workbook()
    ws = wb.active
    ws.append(["ID", "TSA", None])
    ws.append([1, 12530, None])
    ws.append([2, 13000, None])
    ws.merge_cells("B1:C1")
    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


@pytest.fixture()
def merged_data_workbook_bytes() -> bytes:
    """Simulates a customer file where a data cell is merged across multiple rows (e.g. a
    department shared by consecutive employees) - the follower row reads blank without unmerge."""
    wb = Workbook()
    ws = wb.active
    ws.append(["Name", "Department"])
    ws.append(["Alice", "Engineering"])
    ws.append(["Bob", None])
    ws.merge_cells("B2:B3")
    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
