"""Generates template JSON definitions from the real Veryon .xlsx workbooks in app/templates/.

Reads each workbook's header row (from its primary data sheet) and combines it with curated
metadata (id/name/description/required_columns/output styling) to produce the same TemplateDetail
JSON shape the API serves. The source .xlsx files themselves are NOT meant to be committed (see
.gitignore) — this script is how they get turned into the committed JSON definitions.

Run: python scripts/generate_template_definitions.py
"""
import json
from pathlib import Path

import openpyxl

TEMPLATES_DIR = Path(__file__).resolve().parents[1] / "app" / "templates"

# filename -> (preferred source sheet, output id/name/description/sheet_name, required columns,
# header fill color)
TEMPLATE_META: dict[str, dict[str, object]] = {
    "ImportBuildAircraftTemplate-5129 - HT - VALIDATED.xlsx": {
        "source_sheet": "SearchReport",
        "id": "import_build_aircraft",
        "name": "Import Build Aircraft",
        "description": (
            "Aircraft component build/installation records: higher-assembly hierarchy, "
            "times/cycles since new or overhaul, and install/manufacture dates."
        ),
        "sheet_name": "Aircraft Build",
        "required_columns": ["Aircraft", "Model", "Part Number", "Serial Number"],
        "fill": "D9E1F2",
    },
    "Last Done MaintenanceTask EO - 5129.xlsx": {
        "source_sheet": "SearchReport",
        "id": "last_done_maintenance_task",
        "name": "Last Done Maintenance Task",
        "description": (
            "Last-accomplished maintenance task records per component, including "
            "thresholds/intervals and next-due data."
        ),
        "sheet_name": "Last Done Tasks",
        "required_columns": ["SerialNo", "PartNumber", "MaintenanceTask"],
        "fill": "E2EFDA",
    },
    "Maintenance Task - ALL AMP - FIRST.xlsx": {
        "source_sheet": "SearchReport",
        "id": "maintenance_task",
        "name": "Maintenance Task",
        "description": (
            "Maintenance program task definitions (intervals, applicability, and forecasting "
            "flags) for an aircraft maintenance program (AMP)."
        ),
        "sheet_name": "Maintenance Tasks",
        "required_columns": ["MaintenanceTask", "Title"],
        "fill": "FCE4D6",
    },
    "Model - A320 AF MODEL.xlsx": {
        "source_sheet": "SearchReport",
        "id": "model",
        "name": "Model",
        "description": "Aircraft/component model definitions: model type, manufacturer, ICAO model, and applicable aircraft types.",
        "sheet_name": "Models",
        "required_columns": ["ModelType", "Model"],
        "fill": "FFF2CC",
    },
    "Model Part - MINRES.xlsx": {
        "source_sheet": "SearchReport",
        "id": "model_part",
        "name": "Model Part",
        "description": "Part numbers applicable to each model, including default-part flags and serial number ranges.",
        "sheet_name": "Model Parts",
        "required_columns": ["Model", "PartNumber"],
        "fill": "DDEBF7",
    },
    "Model Tree - A320-214-APS 3200 5129.xlsx": {
        "source_sheet": "SearchReport",
        "id": "model_tree",
        "name": "Model Tree",
        "description": "Model hierarchy tree mapping each model to its higher-assembly model, location, and position.",
        "sheet_name": "Model Tree",
        "required_columns": ["Model", "HigherModel"],
        "fill": "F4CCCC",
    },
    "Task Cards - AMP.xlsx": {
        "source_sheet": "SearchReport",
        "id": "task_cards",
        "name": "Task Cards",
        "description": (
            "Maintenance task card (work card) details: man-hours, trade, zone/panel/access "
            "references, and applicability."
        ),
        "sheet_name": "Task Cards",
        "required_columns": ["CardNumber", "Title"],
        "fill": "D9D2E9",
    },
}


def read_headers(path: Path, sheet_name: str) -> list[str]:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb[sheet_name] if sheet_name in wb.sheetnames else wb[wb.sheetnames[0]]
        header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
        return [str(cell) for cell in header_row if cell not in (None, "")]
    finally:
        wb.close()


def column_width(header: str) -> int:
    return max(12, min(40, len(header) + 2))


def main() -> None:
    for filename, meta in TEMPLATE_META.items():
        source_path = TEMPLATES_DIR / filename
        if not source_path.exists():
            print(f"Skipping missing source file: {filename}")
            continue

        columns = read_headers(source_path, str(meta["source_sheet"]))
        template = {
            "id": meta["id"],
            "name": meta["name"],
            "description": meta["description"],
            "sheet_name": meta["sheet_name"],
            "columns": columns,
            "required_columns": meta["required_columns"],
            "output_format": {
                "header_style": {"bold": True, "fill": meta["fill"]},
                "column_widths": {header: column_width(header) for header in columns},
            },
        }

        output_path = TEMPLATES_DIR / f"{meta['id']}.json"
        output_path.write_text(json.dumps(template, indent=2) + "\n", encoding="utf-8")
        print(f"Wrote {output_path} ({len(columns)} columns)")


if __name__ == "__main__":
    main()
