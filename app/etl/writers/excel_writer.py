from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter


def write_excel(
    df: pd.DataFrame, output_path: Path, sheet_name: str, output_format: dict[str, Any]
) -> None:
    """Write a DataFrame to an .xlsx workbook, applying the template's header style/widths."""
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name or "Sheet1"

    header_style = (output_format or {}).get("header_style", {})
    column_widths = (output_format or {}).get("column_widths", {})

    font = Font(bold=bool(header_style.get("bold", False)))
    fill_color = header_style.get("fill")
    fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid") if fill_color else None

    for col_idx, column_name in enumerate(df.columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=column_name)
        cell.font = font
        if fill:
            cell.fill = fill
        width = column_widths.get(column_name)
        if width:
            ws.column_dimensions[get_column_letter(col_idx)].width = width

    for row_idx, row in enumerate(df.itertuples(index=False), start=2):
        for col_idx, value in enumerate(row, start=1):
            ws.cell(row=row_idx, column=col_idx, value=value)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
