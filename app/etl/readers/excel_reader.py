"""Reads customer workbooks via openpyxl directly (no pandas) so we can:

1. Unmerge merged cell ranges and back-fill every cell in the range with the top-left value,
   instead of losing data to blank followers (a common cause of "why is my column empty").
2. Let the caller pick which row is the real header row, instead of always assuming row 1 -
   customer files often have a title/logo row (or several) before the actual headers.
3. Keep blank cells truly blank (Python `None`), never a stray "0"/"nan" placeholder string.
"""
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl.utils import get_column_letter


def _load_unmerged_worksheet(file_path: Path):
    wb = openpyxl.load_workbook(file_path, data_only=True)
    ws = wb[wb.sheetnames[0]]

    # Merged cells only store their value in the top-left cell; every other cell in the range
    # reads as None via iter_rows. Back-fill the value into every cell before unmerging so no
    # data is lost when the sheet is flattened into rows.
    for merged_range in list(ws.merged_cells.ranges):
        min_col, min_row, max_col, max_row = merged_range.bounds
        top_left_value = ws.cell(row=min_row, column=min_col).value
        ws.unmerge_cells(str(merged_range))
        for row in range(min_row, max_row + 1):
            for col in range(min_col, max_col + 1):
                ws.cell(row=row, column=col, value=top_left_value)

    return ws


def _normalize_cell(value: Any) -> Any:
    if value is None:
        return None
    if isinstance(value, str) and value.strip() == "":
        return None
    return value


def _forward_fill_row(row: tuple[Any, ...]) -> list[Any]:
    """Carry the last non-blank value rightward across a header row - higher-order group
    headers (e.g. "Engine 1" spanning the "TSN"/"TSO" columns beneath it) are commonly typed
    once and left blank in the cells they span, rather than actually merged."""
    filled: list[Any] = []
    last_seen: Any = None
    for value in row:
        value = _normalize_cell(value)
        last_seen = value if value is not None else last_seen
        filled.append(last_seen)
    return filled


def _combine_header_levels(header_rows: list[tuple[Any, ...]], num_columns: int) -> list[str]:
    """Combine a range of header rows into one column name per column: every row except the
    bottom-most (the real field-name row) is a higher-order group header, forward-filled and
    chained onto the field name, e.g. "Engine 1 -> TSN". Works for any number of levels."""
    *group_rows, field_row = header_rows
    filled_group_rows = [_forward_fill_row(row) for row in group_rows]

    columns: list[str] = []
    for i in range(num_columns):
        field_value = field_row[i] if i < len(field_row) else None
        field_name = str(field_value).strip() if field_value not in (None, "") else ""

        levels = [
            str(filled[i]).strip()
            for filled in filled_group_rows
            if i < len(filled) and filled[i] not in (None, "")
        ]
        if field_name:
            levels.append(field_name)

        columns.append(" -> ".join(levels) if levels else f"Column{i + 1}")

    return columns


def _dedupe_columns(columns: list[str]) -> tuple[list[str], list[list[int]]]:
    """Collapse columns that share the same header name into one logical column, keeping
    first-seen order. A header cell merged across several columns (e.g. "TSA" spanning J:K) is
    back-filled identically into every cell it spans, so it would otherwise surface as one
    duplicate column per cell it covers - we only want the single "TSA" column back.

    Returns the deduped names plus, for each, the original column indices it was built from.
    """
    index_groups: dict[str, list[int]] = {}
    order: list[str] = []
    for i, name in enumerate(columns):
        if name not in index_groups:
            index_groups[name] = []
            order.append(name)
        index_groups[name].append(i)
    return order, [index_groups[name] for name in order]


def _worksheet_to_rows(
    ws, header_row: int, header_row_start: int | None = None
) -> tuple[list[str], list[dict[str, Any]]]:
    min_row = header_row_start if header_row_start and header_row_start < header_row else header_row
    all_rows = list(ws.iter_rows(min_row=max(min_row, 1), values_only=True))
    if not all_rows:
        return [], []

    header_span = header_row - min_row + 1
    header_rows = all_rows[:header_span]
    data_rows = all_rows[header_span:]

    if len(header_rows) > 1:
        num_columns = max((len(row) for row in header_rows), default=0)
        raw_columns = _combine_header_levels(header_rows, num_columns)
    else:
        raw_columns = []
        for i, cell in enumerate(header_rows[0]):
            name = str(cell).strip() if cell not in (None, "") else ""
            raw_columns.append(name or f"Column{i + 1}")

    columns, index_groups = _dedupe_columns(raw_columns)

    records: list[dict[str, Any]] = []
    for raw_row in data_rows:
        # Skip fully-blank trailing rows rather than emitting a row of Nones.
        if all(_normalize_cell(v) is None for v in raw_row):
            continue
        record: dict[str, Any] = {}
        for name, indices in zip(columns, index_groups):
            value = None
            for idx in indices:
                if idx < len(raw_row):
                    candidate = _normalize_cell(raw_row[idx])
                    if candidate is not None:
                        value = candidate
                        break
            record[name] = value
        records.append(record)

    return columns, records


def read_excel_preview(
    file_path: Path, sample_rows: int = 20, header_row: int = 1, header_row_start: int | None = None
) -> tuple[list[str], list[dict[str, Any]], int]:
    """Return (columns, sample rows, total row count) for an upload preview."""
    ws = _load_unmerged_worksheet(file_path)
    columns, records = _worksheet_to_rows(ws, header_row, header_row_start)
    return columns, records[:sample_rows], len(records)


def read_excel_records(
    file_path: Path, header_row: int = 1, header_row_start: int | None = None
) -> tuple[list[str], list[dict[str, Any]]]:
    """Read the full workbook as (columns, rows) for transformation."""
    ws = _load_unmerged_worksheet(file_path)
    return _worksheet_to_rows(ws, header_row, header_row_start)


def read_raw_grid(file_path: Path, max_rows: int = 200) -> tuple[list[str], list[list[Any]]]:
    """Read the entire sheet as a plain grid (spreadsheet-style "A"/"B"/"C" columns), ignoring
    header_row entirely - used by the constant-value cell picker so the user can pick a value
    from anywhere in the sheet, including rows above/below the chosen header row."""
    ws = _load_unmerged_worksheet(file_path)
    max_col = ws.max_column or 0
    grid_columns = [get_column_letter(i) for i in range(1, max_col + 1)]

    grid_rows: list[list[Any]] = []
    for raw_row in ws.iter_rows(min_row=1, max_row=min(ws.max_row, max_rows), values_only=True):
        grid_rows.append([_normalize_cell(v) for v in raw_row])

    return grid_columns, grid_rows
